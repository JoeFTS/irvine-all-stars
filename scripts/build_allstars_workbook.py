"""Build 2026 All-Stars master workbook from SE registrations CSV with rich fields."""
import csv
import re
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = Path("/Users/joe/irvine-all-stars")
CSV_PATH = ROOT / ".playwright-mcp/unnamed-report.csv"
OUT_PATH = ROOT / "shared-content/documents/2026 All Stars Registration Master.xlsx"

# Map self-declared "Team Name" -> (Tab name, tab color hex AARRGGBB)
TAB_COLOR = {
    "red": "FFFF0000",
    "blue": "FF2F80ED",
    "navy": "FF1A237E",
    "white": "FFEEEEEE",
    "gray": "FF9E9E9E",
    "yellow": "FFFFD60A",
    "shetland": "FF6FCF97",
}

# (csv team name, tab label, color key, full team name, coach name)
TEAM_MAP = [
    ("5u Negrete",        "Negrete 5U",        None,       "Shetland 5U",            "Jose Negrete"),
    ("6u Juarigue",       "Jaurigue 6U Red",   "red",      "Shetland 6U Red",        "Elvin Jaurigue"),
    ("6u Judy",           "Judy Shetland",     "shetland", "Shetland 6U Blue",       "Chad Judy"),
    ("7u MP Singer",      "Singer 7U Red",     "red",      "Pinto MP 7U Red",        "Daniel Singer"),
    ("8u MP Hay",         "Hay 8U Red",        "red",      "Pinto MP 8U Red",        "Brandon Hay"),
    ("8u MP Seto/White",  "WhiteSeto 8U Blue", "blue",     "Pinto MP 8U Blue",       "KC White & Kevin Seto"),
    ("8u KP Watts",       "Watts 8U White",    "white",    "Pinto KP 8U White",      "Randy Watts"),
    ("9u Bernal",         "Bernal 9U Gray",    "gray",     "Mustang 9U Gray",        "Carlos Bernal"),
    ("9u Grifka",         "Grifka 9U Navy",    "navy",     "Mustang 9U Navy",        "Carl Grifka"),
    ("9u Hopp",           "Hopp 9U White",     "white",    "Mustang 9U White",       "Matt Hopp"),
    ("10u Kincade",       "Kincade 10U White", "white",    "Mustang 10U White",      "Ty Kincade"),
    ("10u Stites",        "Stites 10U Gray",   "gray",     "Mustang 10U Gray",       "Robbin Stites & Andy Yang"),
    ("11u Frisch",        "Frisch 11U Red",    "red",      "Bronco 11U Red",         "Dustin Frisch"),
    ("12u Hernandez",     "Hernandez 12U",     None,       "Bronco 12U Red",         "Joe Hernandez"),
    ("12u Sobel/Bailon",  "Bailon 12U Blue",   "blue",     "Bronco 12U Blue",        "Ivan Bailon"),
]


def load_rows():
    rows = []
    with CSV_PATH.open(encoding="utf-8-sig", newline="") as f:
        for r in csv.DictReader(f):
            cleaned = {}
            for k, v in r.items():
                key = k.replace("\\uFEFF", "").replace("﻿", "").strip()
                cleaned[key] = (v or "").strip()
            rows.append(cleaned)
    return rows


def address_one_line(r):
    parts = []
    if r.get("Street Address 1"):
        parts.append(r["Street Address 1"])
    if r.get("Street Address 2"):
        parts.append(r["Street Address 2"])
    csz = []
    if r.get("City"):
        csz.append(r["City"])
    if r.get("State / Province"):
        csz.append(r["State / Province"])
    line2 = ", ".join(csz)
    if r.get("Postal Code"):
        line2 = f"{line2} {r['Postal Code']}".strip()
    out = ", ".join([p for p in parts if p])
    if line2:
        out = f"{out}, {line2}" if out else line2
    return out


def collect_emails(r):
    """Return ordered unique list of all emails for this player."""
    candidates = [
        r.get("Account Email", ""),
        r.get("Email", ""),
        r.get("Email_cp1", ""),
        r.get("Secondary Email", ""),
    ]
    seen, out = set(), []
    for e in candidates:
        e = (e or "").strip().lower()
        if not e or "@" not in e:
            continue
        if e in seen:
            continue
        seen.add(e)
        out.append(e)
    return out


def style_header(ws, headers, fill_color="FF1F4E78", font_color="FFFFFFFF"):
    bold = Font(bold=True, color=font_color)
    fill = PatternFill("solid", fgColor=fill_color)
    thin = Side(style="thin", color="FFCCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    align = Alignment(horizontal="center", vertical="center")
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = bold
        cell.fill = fill
        cell.alignment = align
        cell.border = border


def build_master(wb, rows):
    ws = wb.create_sheet("Master - All Registrations", 0)
    ws.sheet_properties.tabColor = "FF1F77B4"
    headers = [
        "#", "First Name", "Last Name", "Date of Birth", "Team (parent-selected)",
        "Account Email", "Email 2", "Email 3", "Cell Phone", "Address",
        "Registration Date", "Order #", "Status",
    ]
    style_header(ws, headers)
    thin = Side(style="thin", color="FFCCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for i, r in enumerate(rows, start=1):
        emails = collect_emails(r)
        addr = address_one_line(r)
        values = [
            i,
            r.get("First Name", ""),
            r.get("Last Name", ""),
            r.get("Date of Birth", ""),
            r.get("Team Name", ""),
            emails[0] if len(emails) > 0 else "",
            emails[1] if len(emails) > 1 else "",
            emails[2] if len(emails) > 2 else "",
            r.get("Cell Phone Number NEW", ""),
            addr,
            r.get("Registration Date", ""),
            r.get("Order Number", ""),
            r.get("Order Status", ""),
        ]
        for c, v in enumerate(values, 1):
            cell = ws.cell(row=i + 1, column=c, value=v)
            cell.border = border
    widths = [5, 14, 18, 12, 22, 30, 30, 30, 16, 50, 22, 12, 10]
    for c, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(rows) + 1}"


def build_coach_tab(wb, label, color_hex, players, full_team_name, coach_name):
    ws = wb.create_sheet(label)
    if color_hex:
        ws.sheet_properties.tabColor = color_hex
    headers = ["#", "First Name", "Last Name", "Account Email", "Address", "Email 2", "Email 3", "Cell Phone"]
    n_cols = len(headers)
    last_col_letter = get_column_letter(n_cols)

    # Banner row 1: team name, big bold tinted with team color
    banner_fill = color_hex or "FF1F4E78"
    banner_text_color = "FF000000" if color_hex in {TAB_COLOR["white"], TAB_COLOR["yellow"]} else "FFFFFFFF"
    ws.merge_cells(f"A1:{last_col_letter}1")
    b1 = ws.cell(row=1, column=1, value=full_team_name)
    b1.font = Font(bold=True, size=18, color=banner_text_color)
    b1.alignment = Alignment(horizontal="center", vertical="center")
    b1.fill = PatternFill("solid", fgColor=banner_fill)
    ws.row_dimensions[1].height = 30

    # Banner row 2: coach name
    ws.merge_cells(f"A2:{last_col_letter}2")
    b2 = ws.cell(row=2, column=1, value=f"Coach: {coach_name}")
    b2.font = Font(bold=True, size=12, color=banner_text_color)
    b2.alignment = Alignment(horizontal="center", vertical="center")
    b2.fill = PatternFill("solid", fgColor=banner_fill)
    ws.row_dimensions[2].height = 22

    # Headers row 3
    bold = Font(bold=True, color="FFFFFFFF")
    head_fill = PatternFill("solid", fgColor="FF1F4E78")
    thin = Side(style="thin", color="FFCCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    align = Alignment(horizontal="center", vertical="center")
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=c, value=h)
        cell.font = bold
        cell.fill = head_fill
        cell.alignment = align
        cell.border = border

    # Data rows 4+
    for i, r in enumerate(players, start=1):
        emails = collect_emails(r)
        values = [
            i,
            r.get("First Name", ""),
            r.get("Last Name", ""),
            emails[0] if len(emails) > 0 else "",
            address_one_line(r),
            emails[1] if len(emails) > 1 else "",
            emails[2] if len(emails) > 2 else "",
            r.get("Cell Phone Number NEW", ""),
        ]
        for c, v in enumerate(values, 1):
            cell = ws.cell(row=i + 3, column=c, value=v)
            cell.border = border

    widths = [5, 16, 18, 30, 50, 30, 30, 16]
    for c, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:{last_col_letter}{len(players) + 3}"


def main():
    rows = load_rows()
    rows.sort(key=lambda r: (r.get("Last Name", "").lower(), r.get("First Name", "").lower()))
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    build_master(wb, rows)

    grouped = {}
    for r in rows:
        grouped.setdefault(r.get("Team Name", "(unknown)"), []).append(r)

    for csv_team, label, color_key, full_team, coach in TEAM_MAP:
        players = grouped.get(csv_team, [])
        color_hex = TAB_COLOR.get(color_key) if color_key else None
        build_coach_tab(wb, label, color_hex, players, full_team, coach)

    mapped = {csv_team for csv_team, *_ in TEAM_MAP}
    for team, players in grouped.items():
        if team in mapped:
            continue
        build_coach_tab(wb, f"UNMAPPED: {team}"[:31], None, players, team, "?")

    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_PATH)
    print(f"Wrote {OUT_PATH}")
    print(f"Players: {len(rows)} across {len(grouped)} teams")
    for csv_team, label, *_ in TEAM_MAP:
        n = len(grouped.get(csv_team, []))
        print(f"  {label:24} ({csv_team}): {n}")


if __name__ == "__main__":
    main()
