"""Reconcile tryout_registrations parent emails against the master XLSX.

Reads each team sheet in shared-content/documents/2026 All Stars
Registration Master.xlsx and compares each player's primary/secondary
email against the DB row (matched by first+last name within the team).

Prints a discrepancy report. Use --apply to push corrections to DB.
"""

import argparse
import json
import os
import sys
from pathlib import Path
from urllib.error import HTTPError
from urllib.parse import quote
from urllib.request import Request, urlopen

import openpyxl


def load_env(path: Path) -> dict[str, str]:
    env: dict[str, str] = {}
    if not path.exists():
        return env
    for line in path.read_text().splitlines():
        line = line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        k, v = line.split("=", 1)
        env[k.strip()] = v.strip().strip('"').strip("'")
    return env


HERE = Path(__file__).resolve().parent
ENV = {**os.environ, **load_env(HERE.parent / ".env.local")}
SUPA_URL = ENV["NEXT_PUBLIC_SUPABASE_URL"].rstrip("/")
SUPA_KEY = ENV["SUPABASE_SERVICE_ROLE_KEY"]

XLSX = HERE.parent / "shared-content/documents/2026 All Stars Registration Master.xlsx"


def supa(method: str, path: str, body: dict | None = None) -> list[dict]:
    req = Request(
        f"{SUPA_URL}/rest/v1/{path}",
        data=json.dumps(body).encode() if body else None,
        headers={
            "apikey": SUPA_KEY,
            "Authorization": f"Bearer {SUPA_KEY}",
            "Accept-Profile": "irvine_allstars",
            "Content-Profile": "irvine_allstars",
            "Content-Type": "application/json",
            "Prefer": "return=representation",
        },
        method=method,
    )
    try:
        with urlopen(req, timeout=30) as r:
            txt = r.read().decode() or "[]"
            return json.loads(txt) if txt.startswith(("[", "{")) else []
    except HTTPError as e:
        raise RuntimeError(f"{method} {path} -> HTTP {e.code}: {e.read().decode()}")


def norm(s: str | None) -> str:
    return (s or "").strip().lower()


# Master sheet team_name -> DB team_name (where DB differs)
TEAM_ALIAS = {
    "Pinto MP 7U White": "Pinto MP 7U Red",
    "Bronco 11U": "Bronco 11U Red",
}


def name_key(first: str | None, last: str | None) -> str:
    f = norm(first)
    # strip parenthetical nicknames, full-width parens, middle name, "iii"/"jr"
    import re
    f = re.sub(r"[（(].*?[)）]", "", f).strip()
    f = f.split()[0] if f.split() else f  # first token only
    l = norm(last)
    l = re.sub(r"\s+(iii|ii|iv|jr|sr)\.?$", "", l).strip()
    return f"{f}|{l}"


def load_master() -> dict[str, dict]:
    """team_name -> list of rows {first,last,primary,secondary,dob}"""
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    out: dict[str, list[dict]] = {}
    for sheet in wb.sheetnames:
        if sheet == "Master - All Registrations":
            continue
        ws = wb[sheet]
        team_name = ws.cell(1, 1).value
        header_row = None
        for r in range(1, 6):
            if ws.cell(r, 1).value == "#":
                header_row = r
                break
        if not header_row:
            continue
        headers = [ws.cell(header_row, c).value for c in range(1, ws.max_column + 1)]
        rows = []
        for r in range(header_row + 1, ws.max_row + 1):
            rec = {headers[i]: ws.cell(r, i + 1).value for i in range(len(headers))}
            first = rec.get("First Name")
            last = rec.get("Last Name")
            if not first and not last:
                continue
            rows.append({
                "first": (first or "").strip(),
                "last": (last or "").strip(),
                "dob": rec.get("Date of Birth"),
                "primary": norm(rec.get("Account Email")),
                "secondary": norm(rec.get("Email 2")),
            })
        out[team_name] = rows
    return out


def fetch_db() -> list[dict]:
    return supa(
        "GET",
        "tryout_registrations?status=in.(selected,alternate)"
        "&select=id,player_first_name,player_last_name,parent_email,secondary_parent_email,team_id,teams:team_id(team_name)",
    )


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--apply", action="store_true", help="push corrections to DB")
    ap.add_argument("--buckets", default="A,C", help="comma-separated buckets to apply: A,B,C,D")
    ap.add_argument("--exclude-player", action="append", default=[], help="player name substring to skip (case-insensitive); can repeat")
    args = ap.parse_args()

    master = load_master()
    db_rows = fetch_db()

    # index db by team_name + name
    db_index: dict[tuple[str, str], dict] = {}
    for r in db_rows:
        team = (r.get("teams") or {}).get("team_name") if r.get("teams") else None
        if not team:
            continue
        k = (team, name_key(r["player_first_name"], r["player_last_name"]))
        db_index[k] = r

    changes: list[dict] = []
    missing_in_db: list[str] = []
    missing_in_master: list[str] = []

    master_keys: set[tuple[str, str]] = set()
    for team, players in master.items():
        db_team = TEAM_ALIAS.get(team, team)
        for p in players:
            k = (db_team, name_key(p["first"], p["last"]))
            master_keys.add(k)
            db = db_index.get(k)
            if not db:
                missing_in_db.append(f"{db_team}: {p['first']} {p['last']} (in master, not in DB selected/alt)")
                continue
            new_pri = p["primary"] or None
            new_sec = p["secondary"] or None
            cur_pri = norm(db.get("parent_email")) or None
            cur_sec = norm(db.get("secondary_parent_email")) or None
            # treat sec==pri as no secondary
            if cur_sec and cur_sec == cur_pri:
                cur_sec = None
            if new_sec and new_sec == new_pri:
                new_sec = None
            if new_pri and new_pri != cur_pri:
                changes.append({
                    "team": db_team, "player": f"{p['first']} {p['last']}", "id": db["id"],
                    "field": "parent_email", "from": cur_pri, "to": new_pri,
                })
            if new_sec != cur_sec:
                changes.append({
                    "team": db_team, "player": f"{p['first']} {p['last']}", "id": db["id"],
                    "field": "secondary_parent_email", "from": cur_sec, "to": new_sec,
                })

    for k in db_index.keys() - master_keys:
        team, nk = k
        first, last = nk.split("|")
        missing_in_master.append(f"{team}: {first} {last} (in DB selected/alt, not in master sheet)")

    # bucket changes by player id
    by_player: dict[str, dict] = {}
    for c in changes:
        bp = by_player.setdefault(c["id"], {"team": c["team"], "player": c["player"], "fields": {}})
        bp["fields"][c["field"]] = {"from": c["from"], "to": c["to"]}

    bucket_A = []  # master adds secondary DB lacks (pri unchanged)
    bucket_B = []  # DB has secondary master lacks (pri unchanged)
    bucket_C = []  # pri/sec swap (same two emails, different roles)
    bucket_D = []  # primary changed to different email
    for pid, info in by_player.items():
        f = info["fields"]
        pri = f.get("parent_email")
        sec = f.get("secondary_parent_email")
        if pri is None and sec:
            if sec["from"] is None and sec["to"]:
                bucket_A.append((pid, info))
            elif sec["from"] and sec["to"] is None:
                bucket_B.append((pid, info))
            else:
                bucket_D.append((pid, info))  # secondary value changed
        elif pri and sec:
            # check swap: {from_pri, from_sec} == {to_pri, to_sec}
            from_set = {pri["from"], sec["from"]}
            to_set = {pri["to"], sec["to"]}
            if from_set == to_set:
                bucket_C.append((pid, info))
            else:
                bucket_D.append((pid, info))
        elif pri and not sec:
            bucket_D.append((pid, info))
        else:
            bucket_D.append((pid, info))

    def print_bucket(name, items):
        print(f"\n=== {name} ({len(items)}) ===")
        for pid, info in items:
            line = f"  [{info['team']}] {info['player']:28s}"
            for fname, fv in info["fields"].items():
                short = "pri" if fname == "parent_email" else "sec"
                line += f"  {short}: {fv['from']!r}->{fv['to']!r}"
            print(line)

    print_bucket("A. Master ADDS secondary DB lacks (SAFE)", bucket_A)
    print_bucket("B. DB has secondary master lacks (would REMOVE)", bucket_B)
    print_bucket("C. Primary/Secondary swapped (SAFE)", bucket_C)
    print_bucket("D. Primary email different (REVIEW)", bucket_D)

    if missing_in_db:
        print(f"\n=== In master but missing from DB selected/alt ({len(missing_in_db)}) ===")
        for m in missing_in_db:
            print(f"  {m}")

    if missing_in_master:
        print(f"\n=== In DB but missing from master ({len(missing_in_master)}) ===")
        for m in missing_in_master:
            print(f"  {m}")

    if not args.apply:
        print("\n(dry run — pass --apply to push DB updates)")
        return

    selected = set(b.strip().upper() for b in args.buckets.split(","))
    bucket_map = {"A": bucket_A, "B": bucket_B, "C": bucket_C, "D": bucket_D}
    targets: list = []
    for letter, items in bucket_map.items():
        if letter in selected:
            targets.extend(items)

    excludes = [s.lower() for s in args.exclude_player]
    def excluded(name: str) -> bool:
        n = name.lower()
        return any(e in n for e in excludes)

    print(f"\nApplying buckets {sorted(selected)} → {len(targets)} player updates...")
    applied = 0
    for pid, info in targets:
        if excluded(info["player"]):
            print(f"  SKIP (excluded) [{info['team']}] {info['player']}")
            continue
        patch = {fname: fv["to"] for fname, fv in info["fields"].items()}
        supa("PATCH", f"tryout_registrations?id=eq.{pid}", patch)
        print(f"  [{info['team']}] {info['player']} -> {patch}")
        applied += 1
    print(f"done. {applied} applied.")


if __name__ == "__main__":
    main()
